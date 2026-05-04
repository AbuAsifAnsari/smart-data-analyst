# utils/report_generator.py
import io
import re
import os
import tempfile
import pandas as pd
from datetime import datetime


# ─── Helper ───────────────────────────────────────────────────────────────────
def _safe(text: str) -> str:
    if not isinstance(text, str):
        text = str(text)
    text = text.replace("\u2014", "-").replace("\u2013", "-")
    text = text.replace("\u201c", '"').replace("\u201d", '"')
    text = text.replace("\u2018", "'").replace("\u2019", "'")
    text = text.replace("**", "").replace("##", "").replace("#", "")
    text = text.replace("\u2193", "").replace("\u2b07", "")
    return re.sub(r'[^\x00-\x7F]+', '', text).strip()


def _fmt(v):
    try:
        v = float(v)
        if abs(v) >= 1_000_000: return f"{v/1_000_000:.2f}M"
        if abs(v) >= 1_000:     return f"{v:,.0f}"
        return f"{v:.2f}"
    except:
        return str(v)


def _fig_to_png(fig) -> str | None:
    """Plotly figure → temp PNG path. None if fails."""
    try:
        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
            tmp_path = tmp.name
        fig.write_image(tmp_path, format="png", width=680, height=320, scale=1.5)
        return tmp_path
    except Exception as e:
        print(f"[PDF chart error] {e}")
        return None


# ─── Excel Report ─────────────────────────────────────────────────────────────
def generate_excel(df: pd.DataFrame, kpis: dict, messages: list,
                   dataset_name: str) -> bytes:
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        from openpyxl.styles import PatternFill, Font, Alignment

        header_fill = PatternFill("solid", fgColor="3F3FBF")
        header_font = Font(color="FFFFFF", bold=True)
        row_fills   = [PatternFill("solid", fgColor="EEF0FF"),
                       PatternFill("solid", fgColor="FFFFFF")]

        # Sheet 1: KPI Summary
        kpi_df = pd.DataFrame(list(kpis.items()), columns=["Metric", "Value"])
        kpi_df.to_excel(writer, sheet_name="KPI Summary", index=False)
        ws1 = writer.sheets["KPI Summary"]
        ws1.column_dimensions["A"].width = 28
        ws1.column_dimensions["B"].width = 22
        for cell in ws1[1]:
            cell.fill = header_fill; cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for i, row in enumerate(ws1.iter_rows(min_row=2), start=0):
            for cell in row: cell.fill = row_fills[i % 2]

        # Sheet 2: Raw Data
        df.to_excel(writer, sheet_name="Raw Data", index=False)
        ws2 = writer.sheets["Raw Data"]
        for col in ws2.columns:
            max_len = max(len(str(col[0].value or "")),
                          *[len(str(c.value or "")) for c in col[1:]])
            ws2.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
        for cell in ws2[1]:
            cell.fill = header_fill; cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for i, row in enumerate(ws2.iter_rows(min_row=2), start=0):
            for cell in row: cell.fill = row_fills[i % 2]

        # Sheet 3: Chat History
        if messages:
            from utils.ollama_chat import detect_intent, detect_cols, compute_table
            qa_rows = []
            q_num   = 1
            for i, msg in enumerate(messages):
                if msg["role"] == "user":
                    question = msg["content"]
                    answer   = messages[i+1]["content"] if i+1 < len(messages) else ""
                    query    = messages[i+1].get("query","") if i+1 < len(messages) else ""
                    table_str = ""
                    try:
                        intent = detect_intent(question)
                        cols   = detect_cols(question, df)
                        tbl    = compute_table(question, df, intent, cols)
                        if tbl is not None:
                            table_str = "\n\nTable:\n" + tbl.to_string(index=False)
                    except Exception:
                        pass
                    clean = answer.replace("**","").replace("#","").strip()
                    qa_rows.append({
                        "Q No."        : q_num,
                        "Question"     : question,
                        "Answer"       : (clean + table_str)[:3000],
                        "Python Query" : query,
                    })
                    q_num += 1

            if qa_rows:
                chat_df = pd.DataFrame(qa_rows)
                chat_df.to_excel(writer, sheet_name="Chat History", index=False)
                ws3 = writer.sheets["Chat History"]
                ws3.column_dimensions["A"].width = 8
                ws3.column_dimensions["B"].width = 35
                ws3.column_dimensions["C"].width = 70
                ws3.column_dimensions["D"].width = 45
                for cell in ws3[1]:
                    cell.fill = header_fill; cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                for i, row in enumerate(ws3.iter_rows(min_row=2), start=0):
                    for cell in row:
                        cell.fill = row_fills[i % 2]
                        cell.alignment = Alignment(wrap_text=True,
                                                   vertical="top", horizontal="left")
                for row_num, qa in enumerate(qa_rows, start=2):
                    lines = qa["Answer"].count("\n") + 1
                    ws3.row_dimensions[row_num].height = max(40, min(lines*15, 200))

        # Sheet 4: Stats Summary
        numeric_cols = df.select_dtypes(include="number").columns.tolist()
        if numeric_cols:
            summary_df = df[numeric_cols].describe().round(2)
            summary_df.to_excel(writer, sheet_name="Stats Summary")
            ws4 = writer.sheets["Stats Summary"]
            for cell in ws4[1]:
                cell.fill = header_fill; cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

    output.seek(0)
    return output.read()


# ─── PDF Report ───────────────────────────────────────────────────────────────
def generate_pdf(df: pd.DataFrame, kpis: dict, messages: list,
                 dataset_name: str, charts: list = None) -> bytes:
    from fpdf import FPDF

    PAGE_W = 190

    class PDF(FPDF):
        def header(self):
            self.set_font("Helvetica", "B", 14)
            self.set_fill_color(63, 63, 191)
            self.set_text_color(255, 255, 255)
            self.cell(PAGE_W, 12, "Smart Data Analyst - Report",
                      align="C", fill=True, new_x="LMARGIN", new_y="NEXT")
            self.set_text_color(0, 0, 0)
            self.ln(2)

        def footer(self):
            self.set_y(-15)
            self.set_font("Helvetica", "I", 8)
            self.set_text_color(128, 128, 128)
            self.cell(PAGE_W, 10,
                      f"Page {self.page_no()} | Smart Data Analyst", align="C")

        def section_title(self, title: str):
            self.set_font("Helvetica", "B", 12)
            self.set_fill_color(63, 63, 191)
            self.set_text_color(255, 255, 255)
            self.cell(PAGE_W, 9, _safe(title), fill=True,
                      new_x="LMARGIN", new_y="NEXT")
            self.set_text_color(0, 0, 0)
            self.ln(3)

        def subsection(self, title: str):
            self.set_font("Helvetica", "B", 10)
            self.set_fill_color(220, 220, 245)
            self.set_text_color(30, 30, 100)
            self.cell(PAGE_W, 7, _safe(title), fill=True,
                      new_x="LMARGIN", new_y="NEXT")
            self.set_text_color(0, 0, 0)
            self.ln(1)

        def kv_row(self, key: str, value: str, shade: bool = False):
            self.set_fill_color(248, 248, 255) if shade \
                else self.set_fill_color(255, 255, 255)
            self.set_font("Helvetica", "B", 10)
            self.cell(90, 7, _safe(str(key))[:35], border=1, fill=True)
            self.set_font("Helvetica", "", 10)
            self.cell(100, 7, _safe(str(value))[:40], border=1, fill=True,
                      new_x="LMARGIN", new_y="NEXT")

        def safe_cell(self, text: str, h: int = 6,
                      bold: bool = False, fill: bool = False):
            self.set_font("Helvetica", "B" if bold else "", 9)
            self.cell(PAGE_W, h, _safe(str(text))[:95],
                      fill=fill, new_x="LMARGIN", new_y="NEXT")

        def pdf_table(self, tbl_df: pd.DataFrame):
            """DataFrame → proper PDF table with colored header."""
            if tbl_df is None or tbl_df.empty:
                return
            cols  = list(tbl_df.columns)
            n     = len(cols)
            col_w = PAGE_W / n

            # Header
            self.set_font("Helvetica", "B", 8)
            self.set_fill_color(63, 63, 191)
            self.set_text_color(255, 255, 255)
            for col in cols:
                self.cell(col_w, 6, _safe(str(col))[:24],
                          border=1, fill=True, align="C")
            self.ln()
            self.set_text_color(0, 0, 0)

            # Rows
            self.set_font("Helvetica", "", 8)
            row_colors = [(238, 240, 255), (255, 255, 255)]
            for i, (_, row) in enumerate(tbl_df.iterrows()):
                r, g, b = row_colors[i % 2]
                self.set_fill_color(r, g, b)
                for col in cols:
                    val = row[col]
                    val_str = (_fmt(val) if isinstance(val, (int, float))
                               else str(val))
                    align = "R" if isinstance(val, (int, float)) else "L"
                    self.cell(col_w, 6, _safe(val_str)[:24],
                              border=1, fill=True, align=align)
                self.ln()
            self.ln(3)

        def embed_chart(self, fig, caption: str = ""):
            """Plotly figure → PNG → embed in PDF."""
            tmp_path = _fig_to_png(fig)
            if tmp_path:
                if caption:
                    self.set_font("Helvetica", "I", 8)
                    self.set_text_color(80, 80, 80)
                    self.cell(PAGE_W, 5, _safe(caption)[:95],
                              new_x="LMARGIN", new_y="NEXT")
                    self.set_text_color(0, 0, 0)
                self.image(tmp_path, w=PAGE_W)
                self.ln(3)
                try:
                    os.unlink(tmp_path)
                except Exception:
                    pass

    pdf = PDF()
    pdf.set_margins(10, 10, 10)
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Helvetica", "", 10)

    # ── Metadata ──────────────────────────────────────────────────────────────
    pdf.set_text_color(100, 100, 100)
    meta = _safe(
        f"Dataset: {dataset_name}  -  "
        f"Rows: {df.shape[0]:,}  -  Cols: {df.shape[1]}  -  "
        f"Generated: {datetime.now().strftime('%Y-%m-%d')}"
    )[:95]
    pdf.cell(PAGE_W, 6, meta, new_x="LMARGIN", new_y="NEXT")
    pdf.set_text_color(0, 0, 0)
    pdf.ln(4)

    # ── KPI Summary ───────────────────────────────────────────────────────────
    if kpis:
        pdf.section_title("KPI Summary")
        for i, (label, value) in enumerate(kpis.items()):
            pdf.kv_row(label, value, shade=i % 2 == 0)
        pdf.ln(6)

    # ── Dashboard Charts ──────────────────────────────────────────────────────
    if charts:
        pdf.section_title("Dashboard Charts")
        for fig, title in charts:
            pdf.embed_chart(fig, caption=title)

    # ── Data Statistics ───────────────────────────────────────────────────────
    numeric_cols = df.select_dtypes(include="number").columns.tolist()
    if numeric_cols:
        pdf.section_title("Data Statistics")
        for i, col in enumerate(numeric_cols[:6]):
            shade = i % 2 == 0
            pdf.set_fill_color(245, 245, 255) if shade \
                else pdf.set_fill_color(255, 255, 255)
            pdf.safe_cell(f"  {col}", h=6, bold=True, fill=shade)
            pdf.safe_cell(
                f"    Sum: {_fmt(df[col].sum())}   "
                f"Avg: {_fmt(df[col].mean())}   "
                f"Min: {_fmt(df[col].min())}   "
                f"Max: {_fmt(df[col].max())}",
                h=5, fill=shade
            )
        pdf.ln(4)

    # ── Chat Q&A — answer + table + chart per question ────────────────────────
    if messages:
        from utils.ollama_chat import detect_intent, detect_cols, compute_table
        from utils.chart_agent import generate_chart

        pdf.section_title("Chat Q&A Summary")
        q_num = 1

        for i, msg in enumerate(messages):
            if msg["role"] == "user":
                if q_num > 10:
                    break

                question = msg["content"]
                answer   = messages[i+1]["content"] if i+1 < len(messages) else ""
                query    = messages[i+1].get("query","") if i+1 < len(messages) else ""

                # Detect intent + cols
                intent = "general"
                cols   = []
                tbl    = None
                fig    = None
                try:
                    intent = detect_intent(question)
                    cols   = detect_cols(question, df)
                    tbl    = compute_table(question, df, intent, cols)
                    fig    = generate_chart(question, df, intent, cols)
                except Exception:
                    pass

                # ── Q header ─────────────────────────────────────────────────
                pdf.subsection(f"Q{q_num}: {question[:80]}")

                # ── Answer text ───────────────────────────────────────────────
                pdf.set_font("Helvetica", "", 9)
                clean = _safe(answer.replace("**","").replace("#",""))
                clean = re.sub(r'\(table neeche.*?\)', '', clean).strip()
                if clean:
                    for chunk in [clean[j:j+88]
                                  for j in range(0, min(len(clean), 352), 88)]:
                        if chunk.strip():
                            pdf.cell(PAGE_W, 5, chunk,
                                     new_x="LMARGIN", new_y="NEXT")
                    pdf.ln(2)

                # ── Table ─────────────────────────────────────────────────────
                if tbl is not None and not tbl.empty:
                    pdf.pdf_table(tbl.head(10))

                # ── Chart ─────────────────────────────────────────────────────
                if fig is not None:
                    pdf.embed_chart(fig, caption=f"Chart: {question[:60]}")

                # ── Python query (small gray) ─────────────────────────────────
                if query:
                    pdf.set_font("Helvetica", "I", 7)
                    pdf.set_text_color(130, 130, 130)
                    q_clean = _safe(query.replace("\n", " ").strip())
                    pdf.cell(PAGE_W, 5, f"Code: {q_clean[:85]}",
                             new_x="LMARGIN", new_y="NEXT")
                    pdf.set_text_color(0, 0, 0)

                pdf.ln(5)
                q_num += 1

    return bytes(pdf.output())